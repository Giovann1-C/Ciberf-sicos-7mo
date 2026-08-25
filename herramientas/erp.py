# -*- coding: utf-8 -*-
"""ERP del Reto: codigos de material, ordenes de compra (PO), ordenes de
fabricacion (OF) y hojas viajeras. Formato identico a las plantillas del profe.

  python sistema/erp.py estado
  python sistema/erp.py material --categoria MECANICA --tipo Rueda --parte MEC-100 \
         --desc "Rueda mecanum 100mm" --marca Nexus --costo 1800 --lead 3 --sub AGV
  python sistema/erp.py po --proveedor "Nexus Robot" --contacto "Ventas" \
         --email ventas@nexus.com --items "MEC-RUE-MECANUM-100:4,INS-LID-2D-360:1"
  python sistema/erp.py of --parte "AGV Mecanum" --cantidad 1 --entrega 2026-10-18
"""
import os, sys, csv, argparse, datetime, unicodedata

RAIZ   = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
PROY   = os.path.join(RAIZ, "proyecto-reto")
CAT    = os.path.join(PROY, "02-proveeduria", "catalogo-materiales.csv")
ORD    = os.path.join(PROY, "03-compras", "ordenes")
REG_PO = os.path.join(PROY, "03-compras", "registro-po.csv")
OFS    = os.path.join(PROY, "04-fabricacion", "ordenes-fabricacion")
HV     = os.path.join(PROY, "04-fabricacion", "hojas-viajeras")

EMPRESA = "Equipo Reto - Sistemas Ciberfisicos MR3005C.601"
DIRECC  = "Tec de Monterrey, Campus Guadalajara - Av. General Ramon Corona 2514"
MONEDA  = '"$"#,##0.00'
PREFIJO = {"POTENCIA": "POT", "INSTRUMENTACION": "INS", "MECANICA": "MEC",
           "INTERFACES": "INT", "REFACCIONES": "REF", "OTRO": "OTR"}
CAMPOS  = ["codigo_erp", "categoria", "tipo", "numero_parte", "descripcion", "fabricante",
           "uom", "costo_unitario", "stock", "lead_time_semanas", "proveedor",
           "subsistema", "estado"]


def sinacento(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def catalogo():
    if not os.path.exists(CAT):
        return []
    with open(CAT, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def guardar_catalogo(filas):
    with open(CAT, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CAMPOS)
        w.writeheader()
        w.writerows(filas)


def codigo_erp(categoria, tipo, parte):
    p = PREFIJO.get(categoria.upper(), "OTR")
    t = sinacento(tipo).upper().replace(" ", "")[:3]
    n = sinacento(parte).upper().replace(" ", "-")[:12]
    return "%s-%s-%s" % (p, t, n)


def siguiente(registro, prefijo, columna):
    n = 0
    if os.path.exists(registro):
        with open(registro, encoding="utf-8") as f:
            for fila in csv.DictReader(f):
                v = fila.get(columna, "")
                if v.startswith(prefijo):
                    try:
                        n = max(n, int(v.rsplit("-", 1)[1]))
                    except (ValueError, IndexError):
                        pass
    return "%s-%03d" % (prefijo, n + 1)


def lead_max(items):
    return max(int(m["lead_time_semanas"] or 0) for m, _ in items)


# ----------------------------------------------------------------- material
def cmd_material(a):
    filas = catalogo()
    cod = codigo_erp(a.categoria, a.tipo, a.parte)
    if any(f["codigo_erp"] == cod for f in filas):
        print("Ya existe el codigo %s" % cod)
        return 1
    filas.append({"codigo_erp": cod, "categoria": a.categoria.upper(), "tipo": a.tipo,
                  "numero_parte": a.parte, "descripcion": a.desc, "fabricante": a.marca,
                  "uom": a.uom, "costo_unitario": a.costo, "stock": a.stock,
                  "lead_time_semanas": a.lead, "proveedor": a.proveedor,
                  "subsistema": a.sub, "estado": a.estado})
    filas.sort(key=lambda f: f["codigo_erp"])
    guardar_catalogo(filas)

    ficha = os.path.join(PROY, "02-proveeduria", "fichas", cod)
    os.makedirs(ficha, exist_ok=True)
    with open(os.path.join(ficha, "README.md"), "w", encoding="utf-8") as f:
        f.write("# %s\n\n"
                "- **Descripcion:** %s\n- **Parte:** %s\n- **Fabricante:** %s\n"
                "- **Costo unitario:** $%s MXN\n- **Lead time:** %s semanas\n"
                "- **Proveedor:** %s\n- **Subsistema:** %s\n\n"
                "## Ficha tecnica\n_(deja aqui el PDF del datasheet)_\n\n"
                "## Contacto del proveedor\n- Empresa:\n- Atencion:\n- Tel/Email:\n\n"
                "## Cotizaciones\n| Fecha | Proveedor | Precio | Lead time |\n|---|---|---|---|\n"
                % (cod, a.desc, a.parte, a.marca, a.costo, a.lead, a.proveedor, a.sub))
    print("Material creado: %s" % cod)
    print("Ficha: %s" % os.path.relpath(ficha, RAIZ))
    return 0


# ----------------------------------------------------------------------- PO
def cmd_po(a):
    import openpyxl
    from openpyxl.styles import Font, Border, Side, PatternFill

    cat = {f["codigo_erp"]: f for f in catalogo()}
    items = []
    for trozo in a.items.split(","):
        cod, _, cant = trozo.partition(":")
        cod = cod.strip()
        if cod not in cat:
            print("No existe en el catalogo: %s" % cod)
            return 1
        items.append((cat[cod], int(cant or 1)))

    num = siguiente(REG_PO, "OC-2026", "no_orden")
    hoy = datetime.date.today()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orden de Compra"
    neg = Font(bold=True)
    borde = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    relleno = PatternFill("solid", fgColor="D9E1F2")

    ws["A1"] = "ORDEN DE COMPRA"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A4"] = EMPRESA
    ws["A4"].font = neg
    ws["E4"] = "No de Orden:"
    ws["E4"].font = neg
    ws["F4"] = num
    ws["A5"] = DIRECC
    ws["E5"] = "Fecha:"
    ws["E5"].font = neg
    ws["F5"] = hoy.strftime("%d/%m/%Y")
    ws["A6"] = "Proyecto: Celda ciberfisica AGV + Cobot + Vision (CNC Haas)"
    ws["E6"] = "Terminos:"
    ws["E6"].font = neg
    ws["F6"] = a.terminos

    ws["A8"] = "DATOS DEL PROVEEDOR"
    ws["A8"].font = neg
    ws["A8"].fill = relleno
    ws["D8"] = "ENVIAR A"
    ws["D8"].font = neg
    ws["D8"].fill = relleno
    izq = [("Empresa:", a.proveedor), ("Atencion:", a.contacto),
           ("Direccion:", a.direccion), ("Tel/Email:", a.email)]
    der = [("Empresa:", EMPRESA), ("Atencion:", "Rodrigo Herrera"),
           ("Direccion:", DIRECC), ("Notas:", a.notas)]
    for i, (k, v) in enumerate(izq):
        ws.cell(9 + i, 1, k).font = neg
        ws.cell(9 + i, 2, v)
    for i, (k, v) in enumerate(der):
        ws.cell(9 + i, 4, k).font = neg
        ws.cell(9 + i, 5, v)

    enc = ["Item", "Codigo", "Descripcion", "Cantidad", "Unidad", "Precio Unitario", "Importe"]
    for j, e in enumerate(enc, 1):
        c = ws.cell(15, j, e)
        c.font = neg
        c.fill = relleno
        c.border = borde

    total = 0.0
    for i, (m, cant) in enumerate(items, 1):
        precio = float(m["costo_unitario"] or 0)
        imp = precio * cant
        total += imp
        vals = [i, m["codigo_erp"], m["descripcion"], cant, m["uom"], precio, imp]
        for j, v in enumerate(vals, 1):
            ws.cell(15 + i, j, v).border = borde
        ws.cell(15 + i, 6).number_format = MONEDA
        ws.cell(15 + i, 7).number_format = MONEDA

    fin = 16 + len(items)
    ws.cell(fin, 6, "TOTAL MXN").font = neg
    c = ws.cell(fin, 7, total)
    c.font = neg
    c.number_format = MONEDA
    ws.cell(fin + 2, 1, "Lead time mas largo: %d semanas" % lead_max(items)).font = neg
    ws.cell(fin + 3, 1, "Autoriza: ______________________    Recibe: ______________________")
    for col, anch in zip("ABCDEFG", [8, 24, 52, 10, 10, 16, 16]):
        ws.column_dimensions[col].width = anch

    os.makedirs(ORD, exist_ok=True)
    ruta = os.path.join(ORD, "%s_%s.xlsx" % (num, sinacento(a.proveedor).replace(" ", "-")[:25]))
    wb.save(ruta)

    nuevo = not os.path.exists(REG_PO)
    with open(REG_PO, "a", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        if nuevo:
            w.writerow(["no_orden", "fecha", "proveedor", "items", "total_mxn",
                        "lead_max_sem", "estado", "fecha_estimada"])
        lead = lead_max(items)
        w.writerow([num, hoy.isoformat(), a.proveedor, len(items), round(total, 2), lead,
                    "emitida", (hoy + datetime.timedelta(weeks=lead)).isoformat()])

    print("PO generada: %s" % os.path.relpath(ruta, RAIZ))
    print("Total: $%.2f MXN   Lead time: %d semanas" % (total, lead_max(items)))
    print("Registrada en %s" % os.path.relpath(REG_PO, RAIZ))
    return 0


# ----------------------------------------------------------------------- OF
ESTACIONES = [
    (10, "Almacen / Kitting", "Surtido completo del BOM e inspeccion visual de piezas"),
    (20, "Subensamble Mecanico", "Ensamble de chasis y motores. Torque 45 Nm con llave dinamometrica"),
    (30, "Subensamble Electronico", "Ruteo de arneses, paros de emergencia y montaje del controlador"),
    (40, "Integracion de Control", "Carga de firmware NXP, calibracion de encoders, sintonia PI/PD/PID"),
    (50, "QA Electrico", "Continuidad y aislamiento. Prueba de torque estatico"),
    (60, "QA Dinamico", "Validacion LiDAR: freno de emergencia < 0.5 s ante obstaculo a 2 m"),
    (70, "Liberacion", "Firma de calidad y liberacion al cliente"),
]


def cmd_of(a):
    import openpyxl
    from openpyxl.styles import Font, Border, Side, PatternFill

    reg = os.path.join(OFS, "registro-of.csv")
    num = siguiente(reg, "OF-2026-AGV", "no_orden")
    hoy = datetime.date.today()
    neg = Font(bold=True)
    borde = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    relleno = PatternFill("solid", fgColor="D9E1F2")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja Viajera"
    ws["A1"] = "HOJA VIAJERA DE PRODUCCION"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A4"] = "INFORMACION DE LA ORDEN DE TRABAJO"
    ws["A4"].font = neg
    ws["A4"].fill = relleno

    cab = [("No Orden Trabajo:", num, "Fecha Inicio:", hoy.strftime("%d/%m/%Y")),
           ("Producto/Parte:", a.parte, "Fecha Entrega:", a.entrega),
           ("Cantidad Requerida:", a.cantidad, "Cliente:", "Reto MR3005C.601 - Socio Formador")]
    for i, (k, v, k2, v2) in enumerate(cab):
        ws.cell(5 + i, 1, k).font = neg
        ws.cell(5 + i, 2, v)
        ws.cell(5 + i, 4, k2).font = neg
        ws.cell(5 + i, 5, v2)

    enc = ["Op.", "Centro de Trabajo", "Descripcion del Proceso", "Operador",
           "Inicio", "Fin", "Cant. Buena", "Scrap", "Firma Calidad"]
    for j, e in enumerate(enc, 1):
        c = ws.cell(9, j, e)
        c.font = neg
        c.fill = relleno
        c.border = borde
    for i, (op, centro, desc) in enumerate(ESTACIONES, 1):
        for j, v in enumerate([op, centro, desc, "", "", "", "", "", ""], 1):
            ws.cell(9 + i, j, v).border = borde
    for col, anch in zip("ABCDEFGHI", [6, 26, 58, 14, 12, 12, 12, 8, 14]):
        ws.column_dimensions[col].width = anch

    os.makedirs(OFS, exist_ok=True)
    os.makedirs(HV, exist_ok=True)
    ruta = os.path.join(HV, "%s_hoja-viajera.xlsx" % num)
    wb.save(ruta)

    nuevo = not os.path.exists(reg)
    with open(reg, "a", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        if nuevo:
            w.writerow(["no_orden", "fecha_inicio", "parte", "cantidad", "fecha_entrega", "estado"])
        w.writerow([num, hoy.isoformat(), a.parte, a.cantidad, a.entrega, "liberada"])

    print("OF liberada: %s" % num)
    print("Hoja viajera: %s" % os.path.relpath(ruta, RAIZ))
    return 0


# ------------------------------------------------------------------- estado
def cmd_estado(a):
    cat = catalogo()
    print("\n=== CATALOGO DE MATERIALES (%d) ===" % len(cat))
    porsub = {}
    for m in cat:
        porsub.setdefault(m["subsistema"], []).append(m)
    total = 0.0
    for sub in sorted(porsub):
        costo = sum(float(m["costo_unitario"] or 0) for m in porsub[sub])
        total += costo
        print("  %-8s %2d materiales   ~$%10.2f MXN" % (sub, len(porsub[sub]), costo))
    print("  %-8s %2d materiales   ~$%10.2f MXN" % ("TOTAL", len(cat), total))

    hoy = datetime.date.today()
    riesgo = [m for m in cat if int(m["lead_time_semanas"] or 0) >= 5]
    if riesgo:
        print("\n=== RIESGO DE ENTREGA (lead >= 5 semanas) ===")
        for m in sorted(riesgo, key=lambda x: -int(x["lead_time_semanas"])):
            sem = int(m["lead_time_semanas"])
            llega = (hoy + datetime.timedelta(weeks=sem)).strftime("%d/%b")
            print("  %-22s %d sem -> si pides HOY llega %s   [%s]"
                  % (m["codigo_erp"], sem, llega, m["estado"]))

    if os.path.exists(REG_PO):
        with open(REG_PO, encoding="utf-8") as f:
            pos = list(csv.DictReader(f))
        print("\n=== ORDENES DE COMPRA (%d) ===" % len(pos))
        for p in pos:
            print("  %-12s %s  %-26s $%10s  llega %s  [%s]"
                  % (p["no_orden"], p["fecha"], p["proveedor"][:26],
                     p["total_mxn"], p["fecha_estimada"], p["estado"]))
    else:
        print("\n=== ORDENES DE COMPRA ===")
        print("  Ninguna emitida todavia.")

    sinprov = [m for m in cat if m["proveedor"] in ("POR DEFINIR", "")]
    if sinprov:
        print("\n=== SIN PROVEEDOR (%d) - bloquean las compras ===" % len(sinprov))
        print("  " + ", ".join(m["codigo_erp"] for m in sinprov))
    return 0


def main():
    p = argparse.ArgumentParser(description="ERP del Reto")
    sub = p.add_subparsers(dest="cmd", required=True)

    m = sub.add_parser("material")
    m.set_defaults(f=cmd_material)
    m.add_argument("--categoria", required=True, choices=list(PREFIJO))
    m.add_argument("--tipo", required=True)
    m.add_argument("--parte", required=True)
    m.add_argument("--desc", required=True)
    m.add_argument("--marca", default="POR DEFINIR")
    m.add_argument("--uom", default="PZA")
    m.add_argument("--costo", default="0")
    m.add_argument("--stock", default="0")
    m.add_argument("--lead", default="0")
    m.add_argument("--proveedor", default="POR DEFINIR")
    m.add_argument("--sub", default="AGV")
    m.add_argument("--estado", default="por cotizar")

    o = sub.add_parser("po")
    o.set_defaults(f=cmd_po)
    o.add_argument("--proveedor", required=True)
    o.add_argument("--items", required=True)
    o.add_argument("--contacto", default="")
    o.add_argument("--email", default="")
    o.add_argument("--direccion", default="")
    o.add_argument("--terminos", default="30 dias netos")
    o.add_argument("--notas", default="Proyecto academico - entrega en campus")

    f = sub.add_parser("of")
    f.set_defaults(f=cmd_of)
    f.add_argument("--parte", required=True)
    f.add_argument("--cantidad", default="1")
    f.add_argument("--entrega", required=True)

    e = sub.add_parser("estado")
    e.set_defaults(f=cmd_estado)

    a = p.parse_args()
    return a.f(a)


if __name__ == "__main__":
    sys.exit(main())
